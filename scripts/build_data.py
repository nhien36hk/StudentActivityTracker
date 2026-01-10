import sys
import argparse
from pathlib import Path
import pandas as pd

# Thêm parent folder vào sys.path để import src
sys.path.append(str(Path(__file__).resolve().parent.parent))

import json
from openpyxl import load_workbook
from src.extractor import extract_links
from src.downloader import download_file, sanitize_filename
from src.parser import parse_docx_file
from src.sheet_parser import parse_xlsx_file, parse_worksheet
from src.aggregator import aggregate_by_student, save_json, print_summary


# ============ DEFAULT PATHS ============
DEFAULT_EXCEL = Path("data/danhsachct.xlsx")

# Paths (sẽ được set trong main dựa vào arguments)
EXCEL_PATH = DEFAULT_EXCEL
DOWNLOAD_DIR = Path("data/downloaded")
RAW_OUTPUT = Path("data/raw_activities.json")
FINAL_OUTPUT = Path("data/students.json")


def process_external_link(display_text: str, url: str, index: int) -> dict:
    """Download và parse external link (Google Docs/Sheets)."""
    print(f"\n{'='*60}")
    print(f"🌐 [{index}] {display_text[:50]}...")
    
    filename = sanitize_filename(f"{index:03d}_{display_text}")
    file_path = DOWNLOAD_DIR / filename
    
    print(f"⬇️  Downloading...")
    success, file_ext = download_file(url, file_path)
    
    if not success:
        print(f"❌ Download failed!")
        return {'error': 'Download failed', 'url': url, 'students': []}
    
    actual_path = file_path.with_suffix(f'.{file_ext}')
    
    print(f"🔍 Parsing ({file_ext})...")
    if file_ext == 'xlsx':
        activity_name, students = parse_xlsx_file(actual_path, url)
    else:
        activity_name, students = parse_docx_file(actual_path, url)
    
    print(f"✅ {activity_name[:40]}... | {len(students)} sinh viên")
    
    return {
        'activity_name': activity_name,
        'activity_link': url,
        'student_count': len(students),
        'students': students
    }


def process_internal_link(workbook, sheet_name: str, display_text: str, url: str, index: int) -> dict:
    """Parse internal link (sheet trong cùng file Excel)."""
    print(f"\n{'='*60}")
    print(f"📑 [{index}] {display_text[:50]}...")
    
    if sheet_name not in workbook.sheetnames:
        print(f"❌ Sheet không tồn tại: {sheet_name}")
        return {'error': 'Sheet not found', 'students': []}
    
    worksheet = workbook[sheet_name]
    activity_name = sheet_name
    
    print(f"🔍 Parsing sheet: {sheet_name[:40]}...")
    students = parse_worksheet(worksheet, activity_name, activity_link=url)
    
    print(f"✅ {activity_name[:40]}... | {len(students)} sinh viên")
    
    return {
        'activity_name': activity_name,
        'activity_link': url,  # Giữ link gốc
        'student_count': len(students),
        'students': students
    }


def step1_extract_and_parse(limit: int | None) -> list:
    """Bước 1: Extract links và parse (hỗ trợ cả external và internal)."""
    print("\n" + "="*60)
    print("📥 BƯỚC 1: EXTRACT & PARSE")
    print("="*60)
    
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    
    # Extract links
    print(f"\n📂 Đọc file Excel: {EXCEL_PATH}")
    links = extract_links(EXCEL_PATH, limit=limit)
    
    # Count link types
    external_count = sum(1 for l in links if l['link_type'] == 'external')
    internal_count = sum(1 for l in links if l['link_type'] == 'internal')
    
    print(f"✅ Tìm thấy {len(links)} links:")
    print(f"   🌐 External (Google): {external_count}")
    print(f"   📑 Internal (Sheet):  {internal_count}")
    if limit:
        print(f"   (giới hạn {limit})")
    
    # Load workbook một lần cho internal links
    wb = load_workbook(EXCEL_PATH) if internal_count > 0 else None
    
    # Process từng link
    results = []
    for idx, link_info in enumerate(links, 1):
        if link_info['link_type'] == 'external':
            result = process_external_link(
                link_info['display_text'], 
                link_info['url'], 
                idx
            )
        else:
            result = process_internal_link(
                wb,
                link_info['sheet_name'],
                link_info['display_text'],
                link_info['url'],
                idx
            )
        results.append(result)
    
    if wb:
        wb.close()
    
    # Lưu raw data
    print(f"\n💾 Lưu raw data: {RAW_OUTPUT}")
    RAW_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(RAW_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    total_students = sum(r.get('student_count', 0) for r in results)
    print(f"📊 Tổng: {len(results)} chương trình | {total_students} records")
    
    return results


def step2_aggregate(raw_data: list) -> dict:
    """Bước 2: Gom nhóm theo MSSV."""
    print("\n" + "="*60)
    print("🔄 BƯỚC 2: AGGREGATE THEO MSSV")
    print("="*60)
    
    # Flatten students từ raw_data
    all_students = []
    for activity in raw_data:
        all_students.extend(activity.get('students', []))
    
    if not all_students:
        print("⚠️ Không có sinh viên nào!")
        return {}
    
    df = pd.DataFrame(all_students)
    print(f"📋 DataFrame: {len(df)} rows")
    
    result = aggregate_by_student(df)
    
    # Summary
    print_summary(df, result)
    
    # Lưu final data
    print(f"\n💾 Lưu final data: {FINAL_OUTPUT}")
    save_json(result, FINAL_OUTPUT)
    
    return result


def main():
    global EXCEL_PATH, DOWNLOAD_DIR, RAW_OUTPUT, FINAL_OUTPUT
    
    parser = argparse.ArgumentParser(
        description="Build NRL data: Extract → Parse → Aggregate"
    )
    parser.add_argument(
        '-l', '--limit',
        type=int,
        default=None,
        help='Số lượng links xử lý (mặc định: tất cả)'
    )
    parser.add_argument(
        '-e', '--excel',
        type=Path,
        default=DEFAULT_EXCEL,
        help=f'File Excel input (mặc định: {DEFAULT_EXCEL})'
    )
    args = parser.parse_args()
    
    # Set paths dựa vào file Excel
    EXCEL_PATH = args.excel
    if args.excel != DEFAULT_EXCEL:
        excel_name = args.excel.stem  # e.g. "2023-2024"
        DOWNLOAD_DIR = Path(f"data/downloaded_{excel_name}")
        RAW_OUTPUT = Path(f"data/raw_{excel_name}.json")
        FINAL_OUTPUT = Path(f"data/students_{excel_name}.json")
    
    print("🚀 NRL DATA BUILDER")
    print("="*60)
    print(f"   📂 Excel:  {EXCEL_PATH}")
    print(f"   📂 Output: {FINAL_OUTPUT}")
    print(f"   🔢 Limit:  {args.limit if args.limit else 'ALL'}")
    
    # Bước 1: Extract & Parse
    raw_data = step1_extract_and_parse(args.limit)
    
    # Bước 2: Aggregate
    final_data = step2_aggregate(raw_data)
    
    # Hoàn tất
    print("\n" + "="*60)
    print("✅ HOÀN TẤT!")
    print("="*60)
    print(f"   📁 Raw data:   {RAW_OUTPUT}")
    print(f"   📁 Final data: {FINAL_OUTPUT}")
    print(f"   👥 Sinh viên:  {len(final_data)}")


if __name__ == "__main__":
    main()
