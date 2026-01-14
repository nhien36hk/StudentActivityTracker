"""
Module extract hyperlinks từ file Excel.

Logic đơn giản:
- File năm ≤ 2022: Tất cả links là internal (sheet trong cùng file)
- File năm > 2022: Có thể có external links (Google Docs/Sheets)
"""
from pathlib import Path
from typing import List, Tuple, Dict, Any
from openpyxl import load_workbook
import re

MASTER_LINK = {
    2020: 'https://docs.google.com/spreadsheets/d/1QWdhplM8SzatAbQUG5N8xlefUCZGEIfZ/edit?gid=1391010033#gid=1391010033',
    2021: None,
    2022: 'https://docs.google.com/spreadsheets/d/1fWDcSwa9p3lbOceJOaBluFqGT9JPPwiF/edit?gid=1627022443#gid=1627022443',
}


def get_file_year(excel_path: Path) -> int:
    """Lấy năm từ tên file. VD: '2022-2023.xlsx' → 2022"""
    match = re.search(r'(\d{4})', excel_path.stem)
    return int(match.group(1)) if match else 9999


def find_link_column(worksheet, header_row: int = 2) -> int | None:
    """Tìm cột chứa link trong worksheet. Hỗ trợ: 'Link', 'Sheet'."""
    keywords = ['link', 'sheet', 'quyết định công nhận nrl']
    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col).value
        if cell_value:
            cell_lower = str(cell_value).lower()
            if any(kw in cell_lower for kw in keywords):
                return col
    return None


def parse_sheet_location(location: str) -> str | None:
    """
    Parse tên sheet từ internal link location.
    - "'TÊN SHEET'!A1" → "TÊN SHEET"
    - "TÊN SHEET!A1" → "TÊN SHEET"
    """
    if not location:
        return None
    
    # Format có quotes
    match = re.match(r"'(.+?)'!", location)
    if match:
        return match.group(1)
    
    # Format không quotes
    match = re.match(r"(.+?)!\$?[A-Z]+\$?\d+", location)
    if match:
        return match.group(1).strip()
    
    return None


def extract_links(excel_path: Path, limit: int | None = None) -> List[Dict[str, Any]]:
    """
    Extract danh sách links từ file Excel.
    
    - File năm ≤ 2022: Tất cả là internal sheet
    - File năm > 2022: Check external/internal
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    
    link_col = find_link_column(ws)
    if not link_col:
        raise ValueError("Không tìm thấy cột Link trong file Excel")
    
    file_year = get_file_year(excel_path)
    is_old_file = file_year <= 2022  # File cũ = tất cả internal
    
    links: List[Dict[str, Any]] = []
    header_row = 2
    
    for row in range(header_row + 1, ws.max_row + 1):
        if limit and len(links) >= limit:
            break
            
        cell = ws.cell(row=row, column=link_col)
        if not cell.hyperlink:
            continue
        
        display_text = str(cell.value) if cell.value else ""
        location = cell.hyperlink.location
        
        # File cũ (≤2022): Tất cả là internal sheet
        if is_old_file and location:
            sheet_name = parse_sheet_location(location)
            if sheet_name and sheet_name in wb.sheetnames:
                links.append({
                    'display_text': display_text,
                    'link_type': 'internal',
                    'url': MASTER_LINK[file_year] if not file_year == 2021 else display_text,
                    'sheet_name': sheet_name,
                })
        
        # File mới (>2022): Check external trước
        elif not is_old_file:
            if cell.hyperlink.target:
                links.append({
                    'display_text': display_text,
                    'link_type': 'external',
                    'url': cell.hyperlink.target,
                    'sheet_name': None,
                })
            elif location:
                sheet_name = parse_sheet_location(location)
                if sheet_name and sheet_name in wb.sheetnames:
                    url = display_text if display_text.startswith('http') else location
                    links.append({
                        'display_text': display_text,
                        'link_type': 'internal',
                        'url': url,
                        'sheet_name': sheet_name,
                    })
    
    wb.close()
    return links


# Legacy function
def extract_hyperlinks(excel_path: Path, limit: int | None = None) -> List[Tuple[str, str]]:
    """Legacy: Chỉ trả về external links."""
    links = extract_links(excel_path, limit)
    return [(l['display_text'], l['url']) for l in links if l['link_type'] == 'external']
