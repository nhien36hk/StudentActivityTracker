"""
Module parse file Excel (Google Sheets) để trích xuất thông tin sinh viên.
Hỗ trợ parse từ file hoặc từ worksheet trực tiếp.
"""
from pathlib import Path
from typing import List, Dict, Tuple
import re
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    """
    Xác định tên cột trong DataFrame.
    Hỗ trợ nhiều biến thể tên cột: "Họ và tên", "Họ tên", "Ho ten", v.v.
    
    Returns:
        Dict mapping role -> column_name
    """
    columns = {}
    
    for col in df.columns:
        col_lower = str(col).lower().strip()
        
        if 'stt' in col_lower:
            columns['stt'] = col
        elif 'mssv' in col_lower or 'mã sv' in col_lower or 'mã sinh viên' in col_lower:
            columns['student_id'] = col
        elif ('họ' in col_lower and 'tên' in col_lower) or 'ho ten' in col_lower or 'hoten' in col_lower:
            # Matches: "Họ và tên", "Họ tên", "Ho ten", "Họ và tên:", etc.
            columns['name'] = col
        elif 'lớp' in col_lower or 'đơn vị' in col_lower:
            columns['student_class'] = col
        elif 'nrl' in col_lower or 'điểm' in col_lower:
            columns['score'] = col
    
    return columns


def clean_student_id(raw_id) -> str:
    """Làm sạch MSSV."""
    if pd.isna(raw_id):
        return ""
    return str(raw_id).strip().upper().replace(" ", "").replace(".0", "")


def is_student_id(value: str) -> bool:
    """Kiểm tra có phải MSSV không."""
    clean_val = value.strip().replace(" ", "")
    return clean_val.isdigit() and len(clean_val) >= 8


def is_class_code(value: str) -> bool:
    """Kiểm tra có phải mã lớp không."""
    clean_val = value.strip().replace(' ', '').upper()
    return bool(re.match(r'^\d{2}[A-ZĐÀÁẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬ]+\d*$', clean_val))


def smart_swap_id_class(raw_id: str, raw_class: str) -> Tuple[str, str]:
    """Swap student_id/student_class nếu bị đảo."""
    raw_id = str(raw_id).strip() if raw_id else ""
    raw_class = str(raw_class).strip() if raw_class else ""
    
    if is_class_code(raw_id) and is_student_id(raw_class):
        return raw_class, raw_id
    if not raw_id and is_student_id(raw_class):
        return raw_class, ""
    if is_class_code(raw_id) and not raw_class:
        return "", raw_id
    
    return raw_id, raw_class


def parse_score(score_val) -> float:
    """Parse điểm NRL. Bỏ qua datetime và giá trị bất thường."""
    from datetime import datetime
    
    if pd.isna(score_val):
        return 0.0
    
    # Skip datetime (lỗi data trong Excel)
    if isinstance(score_val, datetime):
        return 0.0
    
    clean_text = str(score_val).strip().replace(',', '.')
    match = re.search(r'(\d+\.?\d*)', clean_text)
    if match:
        try:
            score = float(match.group(1))
            # Điểm NRL hợp lệ thường < 100, nếu > 1000 có thể là năm
            return score if score < 1000 else 0.0
        except ValueError:
            return 0.0
    return 0.0


def extract_activity_name(file_path: Path) -> str:
    """Trích xuất tên chương trình từ tên file."""
    filename = file_path.stem
    name = re.sub(r'^\d+_', '', filename)
    name = re.sub(r'^QĐxx\d+\s*-\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'^(CÔNG NHẬN|Công nhận)\s+NRL\s*', '', name, flags=re.IGNORECASE)
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    return name if name else "Unknown"


def parse_xlsx_file(file_path: Path, activity_link: str) -> Tuple[str, List[Dict]]:
    """
    Parse file Excel, trả về tên chương trình và danh sách sinh viên.
    
    Args:
        file_path: Đường dẫn file .xlsx
        activity_link: Link gốc của chương trình
        
    Returns:
        Tuple (tên chương trình, list sinh viên)
    """
    activity_name = extract_activity_name(file_path)
    
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"❌ Lỗi đọc Excel: {e}")
        return activity_name, []
    
    if df.empty:
        return activity_name, []
    
    # Detect columns
    columns = detect_columns(df)
    
    if 'name' not in columns:
        print(f"⚠️ Không tìm thấy cột Họ tên trong file")
        return activity_name, []
    
    students = []
    
    for idx, row in df.iterrows():
        # Lấy giá trị
        stt = int(row.get(columns.get('stt', ''), idx + 1) or idx + 1)
        name = str(row.get(columns.get('name', ''), '')).strip()
        raw_id = row.get(columns.get('student_id', ''), '')
        raw_class = row.get(columns.get('student_class', ''), '')
        score_val = row.get(columns.get('score', ''), 0)
        
        # Smart swap
        student_id, student_class = smart_swap_id_class(raw_id, raw_class)
        student_id = clean_student_id(student_id)
        student_class = str(student_class).strip() if student_class else ""
        
        # Skip empty rows
        if not student_id and not name:
            continue
        
        # Mark unknown
        if not name:
            name = "UNKNOWN_NAME"
        if not student_id:
            student_id = f"UNKNOWN_ID_{stt}"
        if not student_class:
            student_class = "UNKNOWN_CLASS"
        
        students.append({
            'stt': stt,
            'name': name,
            'student_id': student_id,
            'student_class': student_class,
            'score': parse_score(score_val),
            'activity_name': activity_name,
            'activity_link': activity_link,
        })
    
    return activity_name, students


def find_header_row(worksheet: Worksheet) -> int:
    """
    Tìm dòng header trong worksheet.
    Header thường chứa "Họ" và "tên" hoặc "MSSV".
    
    Returns:
        Số dòng header (1-indexed), mặc định là 2
    """
    for row in range(1, min(worksheet.max_row + 1, 10)):
        row_text = " ".join(str(worksheet.cell(row=row, column=c).value or "").lower() 
                          for c in range(1, min(worksheet.max_column + 1, 15)))
        if ('họ' in row_text and 'tên' in row_text) or 'mssv' in row_text:
            return row
    return 2  # Default


def parse_worksheet(worksheet: Worksheet, activity_name: str, activity_link: str = "") -> List[Dict]:
    """
    Parse worksheet trực tiếp từ openpyxl worksheet object.
    Tự động detect header row (thường ở row 2).
    
    Args:
        worksheet: Worksheet object từ openpyxl
        activity_name: Tên chương trình
        activity_link: Link (có thể để trống cho internal)
        
    Returns:
        List sinh viên
    """
    if worksheet.max_row < 2:
        return []
    
    # Tìm header row
    header_row = find_header_row(worksheet)
    
    # Lấy header
    headers = [worksheet.cell(row=header_row, column=c).value for c in range(1, worksheet.max_column + 1)]
    
    # Lấy data từ row sau header
    data_rows = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        row_data = [worksheet.cell(row=row, column=c).value for c in range(1, worksheet.max_column + 1)]
        data_rows.append(row_data)
    
    if not data_rows:
        return []
    
    # Xử lý cột trùng tên: thêm suffix để unique
    unique_headers = []
    seen = {}
    for h in headers:
        h_str = str(h) if h else 'None'
        if h_str in seen:
            seen[h_str] += 1
            unique_headers.append(f"{h_str}_{seen[h_str]}")
        else:
            seen[h_str] = 0
            unique_headers.append(h_str)
    
    # Tạo DataFrame với headers unique
    df = pd.DataFrame(data_rows, columns=unique_headers)
    
    if df.empty:
        return []
    
    # Detect columns (dùng headers gốc để detect, nhưng map sang unique)
    # Tạo df tạm với headers gốc để detect
    df_temp = pd.DataFrame(data_rows, columns=[str(h) if h else 'None' for h in headers])
    cols = detect_columns(df_temp)
    
    # Map column names sang unique headers
    for key in cols:
        original = cols[key]
        if original in unique_headers:
            continue
        # Tìm index của cột gốc
        try:
            idx = [str(h) if h else 'None' for h in headers].index(str(original))
            cols[key] = unique_headers[idx]
        except ValueError:
            pass
    
    if 'name' not in cols:
        print(f"⚠️ Không tìm thấy cột Họ tên trong sheet (header row {header_row})")
        return []
    
    students = []
    
    def safe_get(row, col_name, default=None):
        """Lấy giá trị từ row, xử lý trường hợp Series (cột trùng tên)."""
        val = row.get(col_name, default)
        # Nếu là Series (do cột trùng tên), lấy giá trị đầu tiên
        if isinstance(val, pd.Series):
            val = val.iloc[0] if len(val) > 0 else default
        return val
    
    for idx, row in df.iterrows():
        # Lấy STT
        stt_val = safe_get(row, cols.get('stt', ''), None)
        try:
            stt = int(stt_val) if pd.notna(stt_val) else idx + 1
        except (ValueError, TypeError):
            stt = idx + 1
        
        name_val = safe_get(row, cols.get('name', ''), '')
        name = str(name_val).strip() if pd.notna(name_val) else ""
        
        raw_id = safe_get(row, cols.get('student_id', ''), '')
        raw_class = safe_get(row, cols.get('student_class', ''), '')
        score_val = safe_get(row, cols.get('score', ''), 0)
        
        # Smart swap
        student_id, student_class = smart_swap_id_class(raw_id, raw_class)
        student_id = clean_student_id(student_id)
        student_class = str(student_class).strip() if pd.notna(student_class) else ""
        
        # Check invalid values
        invalid_ids = ['', 'NAN', 'NONE', 'NULL']
        invalid_names = ['', 'nan', 'none', 'null']
        
        id_is_invalid = not student_id or student_id in invalid_ids
        name_is_invalid = not name or name.lower() in invalid_names
        
        # Skip nếu CẢ HAI đều invalid
        if id_is_invalid and name_is_invalid:
            continue
        
        # Mark unknown (chỉ khi 1 trong 2 có giá trị)
        if name_is_invalid:
            name = "UNKNOWN_NAME"
        if id_is_invalid:
            student_id = f"UNKNOWN_ID_{stt}"
        if not student_class or student_class.lower() in ['nan', 'none', 'null']:
            student_class = "UNKNOWN_CLASS"
        
        students.append({
            'stt': stt,
            'name': name,
            'student_id': student_id,
            'student_class': student_class,
            'score': parse_score(score_val),
            'activity_name': activity_name,
            'activity_link': activity_link,
        })
    
    return students

